import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

# 创建主窗口
root = tk.Tk()
root.title("任务管理程序")

# 打开Excel文件的函数
def open_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        load_tasks_from_excel(file_path)

# 从Excel文件中加载任务
def load_tasks_from_excel(file_path):
    # 清空已有的按钮
    for widget in frame.winfo_children():
        widget.destroy()

    # 打开Excel文件
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 读取数据并创建按钮
    for row in sheet.iter_rows(min_row=2, values_only=True):
        task_name, task_duration, start_time, end_time = row
        print(row)
        button = tk.Button(frame, text=task_name, width=task_duration * 5)
        button.grid(row=0, column=0, padx=(start_time.hour * 60 + start_time.minute), pady=10)

# 创建打开文件按钮
open_button = tk.Button(root, text="打开Excel文件", command=open_excel_file)
open_button.pack(pady=10)

# 创建任务按钮的框架
frame = tk.Frame(root)
frame.pack()

root.mainloop()
